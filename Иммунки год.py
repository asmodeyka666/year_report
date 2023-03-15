from tkinter import filedialog as fd
import tkinter as tk
import tkinter.messagebox as mb
import pandas as pd
import os
import openpyxl as ox
import os.path


def creat_sheet(wb, file_name_dow):
    sheet_xls = wb[wb.sheetnames[-2]]
    wb.copy_worksheet(sheet_xls)
    file_date = os.path.basename(file_name_dow)[0:10].replace('_', '.')
    name_list = file_date
    if name_list in wb.sheetnames:
        name_list = name_list + ' new'
    wb[wb.sheetnames[-1]].title = name_list
    wb.move_sheet(name_list, offset=-1)
    return name_list


def diphtheria(df, file_name_dow):
    for file in files_report:
        if 'дифтерия' in os.path.basename(file).lower():
            wb = ox.load_workbook(filename=file, read_only=False)
            break
    name_list = creat_sheet(wb, file_name_dow)
    ws = wb[name_list]
    df_short = df.iloc[0:21, :]
    for ir in range(0, len(df_short)):
        for ic in range(0, len(df_short.iloc[ir])):
            ws.cell(6 + ir, 1 + ic).value = df_short.iloc[ir][ic]
    wb.save(file)
    # заполняем лист Динамика показателей
    ws = wb['динамика показателей']
    # Ищем последнюю заполненную строку
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == None:
            last_row = row
            break
    col_dict = {2: 'G29', 5: 'G34', 8: 'Y28', 11: 'AA28', 14: 'U28', 18: 'B', 19: 'E', 23: 'N', 24: 'H', 25: 'K'}
    for col in range(1, 26):
        if col in [1, 4, 7, 10, 13, 17, 22]:
            ws.cell(last_row, col).value = name_list
        elif col < 17 and col in col_dict:
            ws.cell(last_row, col).value = f"='{name_list}'!{col_dict[col]}"
        elif col > 17 and col in col_dict:
            ws.cell(last_row, col).value = f"={col_dict[col]}{last_row}"
    wb.save(file)
    wb.close()


def pertussis(df, file_name_dow):
    for file in files_report:
        if 'коклюш' in os.path.basename(file).lower():
            wb = ox.load_workbook(filename=file, read_only=False)
            break
    name_list = creat_sheet(wb, file_name_dow)
    ws = wb[name_list]
    df_short = df.iloc[0:21, :]
    for ir in range(0, len(df_short)):
        for ic in range(0, len(df_short.iloc[ir])):
            ws.cell(6 + ir, 1 + ic).value = df_short.iloc[ir][ic]
    wb.save(file)
    # заполняем лист Динамика показателей
    ws = wb['динамика показателей']
    # Ищем последнюю заполненную строку
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == None:
            last_row = row
            break
    col_dict = {2: 'G32', 5: 'G37', 8: 'W29', 11: 'X29', 14: 'U29', 19: 'B', 20: 'E', 24: 'N', 25: 'H', 26: 'K'}
    for col in range(1, 27):
        if col in [1, 4, 7, 10, 13, 18, 23]:
            ws.cell(last_row, col).value = name_list
        elif col < 17 and col in col_dict:
            ws.cell(last_row, col).value = f"='{name_list}'!{col_dict[col]}"
        elif col > 17 and col in col_dict:
            ws.cell(last_row, col).value = f"={col_dict[col]}{last_row}"
    wb.save(file)
    wb.close()


def measles(df, file_name_dow):
    for file in files_report:
        if 'корь' in os.path.basename(file).lower():
            wb = ox.load_workbook(filename=file, read_only=False)
            break
    name_list = creat_sheet(wb, file_name_dow)
    ws = wb[name_list]
    df_short = df.iloc[0:19, :]
    for ir in range(0, len(df_short)):
        for ic in range(0, len(df_short.iloc[ir])):
            ws.cell(6 + ir, 1 + ic).value = df_short.iloc[ir][ic]
    wb.save(file)
    # заполняем лист Динамика показателей
    ws = wb['динамика показателей']
    # Ищем последнюю заполненную строку
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == None:
            last_row = row
            break
    col_dict = {2: 'G30', 5: 'G35', 8: 'Q27', 11: 'S27', 14: 'M27', 19: 'B', 20: 'E', 24: 'N', 25: 'H', 26: 'K'}
    for col in range(1, 27):
        if col in [1, 4, 7, 10, 13, 18, 23]:
            ws.cell(last_row, col).value = name_list
        elif col < 17 and col in col_dict:
            ws.cell(last_row, col).value = f"='{name_list}'!{col_dict[col]}"
        elif col > 17 and col in col_dict:
            ws.cell(last_row, col).value = f"={col_dict[col]}{last_row}"
    wb.save(file)
    wb.close()


def rubella(df, file_name_dow):
    for file in files_report:
        if 'краснуха' in os.path.basename(file).lower():
            wb = ox.load_workbook(filename=file, read_only=False)
            break
    name_list = creat_sheet(wb, file_name_dow)
    ws = wb[name_list]
    df_short = df.iloc[0:19, :]
    for ir in range(0, len(df_short)):
        for ic in range(0, len(df_short.iloc[ir])):
            ws.cell(6 + ir, 1 + ic).value = df_short.iloc[ir][ic]
    wb.save(file)
    # заполняем лист Динамика показателей
    ws = wb['динамика показателей']
    # Ищем последнюю заполненную строку
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == None:
            last_row = row
            break
    col_dict = {2: 'G30', 5: 'G35', 8: 'O27', 11: 'Q27', 14: 'K27', 19: 'B', 20: 'E', 24: 'N', 25: 'H', 26: 'K'}
    for col in range(1, 27):
        if col in [1, 4, 7, 10, 13, 18, 23]:
            ws.cell(last_row, col).value = name_list
        elif col < 17 and col in col_dict:
            ws.cell(last_row, col).value = f"='{name_list}'!{col_dict[col]}"
        elif col > 17 and col in col_dict:
            ws.cell(last_row, col).value = f"={col_dict[col]}{last_row}"
    wb.save(file)
    wb.close()


def mumps(df, file_name_dow):
    for file in files_report:
        if 'паротит' in os.path.basename(file).lower():
            wb = ox.load_workbook(filename=file, read_only=False)
            break
    name_list = creat_sheet(wb, file_name_dow)
    ws = wb[name_list]
    df_short = df.iloc[0:19, :]
    for ir in range(0, len(df_short)):
        for ic in range(0, len(df_short.iloc[ir])):
            ws.cell(6 + ir, 1 + ic).value = df_short.iloc[ir][ic]
    wb.save(file)
    # заполняем лист Динамика показателей
    ws = wb['динамика показателей']
    # Ищем последнюю заполненную строку
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == None:
            last_row = row
            break
    col_dict = {2: 'G30', 5: 'G35', 8: 'O27', 11: 'Q27', 14: 'K27', 19: 'B', 20: 'E', 24: 'N', 25: 'H', 26: 'K'}
    for col in range(1, 27):
        if col in [1, 4, 7, 10, 13, 18, 23]:
            ws.cell(last_row, col).value = name_list
        elif col < 17 and col in col_dict:
            ws.cell(last_row, col).value = f"='{name_list}'!{col_dict[col]}"
        elif col > 17 and col in col_dict:
            ws.cell(last_row, col).value = f"={col_dict[col]}{last_row}"
    wb.save(file)
    wb.close()


def poliomyelitis(df, file_name_dow):
    for file in files_report:
        if 'полиомиелит' in os.path.basename(file).lower():
            wb = ox.load_workbook(filename=file, read_only=False)
            break
    name_list = creat_sheet(wb, file_name_dow)
    ws = wb[name_list]
    df_short = df.iloc[0:10, :]
    for ir in range(0, len(df_short)):
        for ic in range(0, len(df_short.iloc[ir])):
            ws.cell(6 + ir, 1 + ic).value = df_short.iloc[ir][ic]

    # заполняем лист Динамика показателей
    ws = wb['динамика показателей']
    # Ищем последнюю заполненную строку
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == None:
            last_row = row
            break
    col_dict = {2: 'G32', 5: 'G37', 8: 'S29', 11: 'T29', 14: 'Q29', 18: 'B', 19: 'E', 23: 'N', 24: 'H', 25: 'K'}
    for col in range(1, 26):
        if col in [1, 4, 7, 10, 13, 17, 22]:
            ws.cell(last_row, col).value = name_list
        elif col < 17 and col in col_dict:
            ws.cell(last_row, col).value = f"='{name_list}'!{col_dict[col]}"
        elif col > 17 and col in col_dict:
            ws.cell(last_row, col).value = f"={col_dict[col]}{last_row}"
    wb.save(file)
    wb.close()


def tetanus(df, file_name_dow):
    for file in files_report:
        if 'столбняк' in os.path.basename(file).lower():
            wb = ox.load_workbook(filename=file, read_only=False)
            break
    name_list = creat_sheet(wb, file_name_dow)
    ws = wb[name_list]
    df_short = df.iloc[0:21, :]
    for ir in range(0, len(df_short)):
        for ic in range(0, len(df_short.iloc[ir])):
            ws.cell(6 + ir, 1 + ic).value = df_short.iloc[ir][ic]

    # заполняем лист Динамика показателей
    ws = wb['динамика показателей']
    # Ищем последнюю заполненную строку
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == None:
            last_row = row
            break
    col_dict = {2: 'G29', 5: 'G34', 8: 'Y28', 11: 'AA28', 14: 'U28', 18: 'B', 19: 'E', 23: 'N', 24: 'H', 25: 'K'}
    for col in range(1, 26):
        if col in [1, 4, 7, 10, 13, 17, 22]:
            ws.cell(last_row, col).value = name_list
        elif col < 17 and col in col_dict:
            ws.cell(last_row, col).value = f"='{name_list}'!{col_dict[col]}"
        elif col > 17 and col in col_dict:
            ws.cell(last_row, col).value = f"={col_dict[col]}{last_row}"
    wb.save(file)
    wb.close()


def tuberculosis(df, file_name_dow):
    for file in files_report:
        if 'туберкулез' in os.path.basename(file).lower():
            wb = ox.load_workbook(filename=file, read_only=False)
            break
    name_list = creat_sheet(wb, file_name_dow)
    ws = wb[name_list]
    df_short = df.iloc[0:16, :]
    for ir in range(0, len(df_short)):
        for ic in range(0, len(df_short.iloc[ir])):
            ws.cell(6 + ir, 1 + ic).value = df_short.iloc[ir][ic]

    # заполняем лист Динамика показателей
    ws = wb['динамика показателей']
    # Ищем последнюю заполненную строку
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == None:
            last_row = row
            break
    col_dict = {2: 'G27', 5: 'J24', 8: 'K24', 11: 'I24', 18: 'K', 19: 'E', 20: 'H', 21: 'B'}
    for col in range(1, 22):
        if col in [1, 4, 7, 10, 17]:
            ws.cell(last_row, col).value = name_list
        elif col < 17 and col in col_dict:
            ws.cell(last_row, col).value = f"='{name_list}'!{col_dict[col]}"
        elif col > 17 and col in col_dict:
            ws.cell(last_row, col).value = f"={col_dict[col]}{last_row}"
    wb.save(file)
    wb.close()


functions = {
    'дифтерия': diphtheria,
    'коклюш': pertussis,
    'корь': measles,
    'краснуха': rubella,
    'паротит': mumps,
    'полиомиелит': poliomyelitis,
    'столбняк': tetanus,
    'туберкулез': tuberculosis
}
root = tk.Tk()
root.withdraw()
files_report = fd.askopenfilename(multiple=True, title='Выберите файлы с иммунками ГОД')
files_dowload = fd.askopenfilename(multiple=True, title='Выберите файлы со СВОДОМ из VACC_VACC')
files_report = list(files_report)

# проверка комплиментарности инфекций в отчетах и загруженных файлах
inf_rep = []
for rep in files_report:
    inf_rep.append(os.path.basename(rep).split(' ')[0].lower())
for dow in files_dowload:
    inf = os.path.basename(dow).split('_')[-2].lower()
    if inf not in inf_rep:
        mb.showerror(f'Не выбран файл отчета для {inf}',
                     f'Не выбран файл отчета для {inf.upper()} \n Выберите нужный отчет')
        files_report.append(fd.askopenfilename(title=f'Выберите файл с иммунками ГОД для {inf.upper()}'))

# загружаем данные выгрузок
for dow in files_dowload:
    df = pd.read_excel(dow, skiprows=6, header=None)
    inf = os.path.basename(dow).split('_')[-2].lower()
    functions[inf](df, dow)
